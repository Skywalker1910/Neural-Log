from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from datetime import datetime, timedelta
import sqlite3
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'
# Database migration completed - is_admin column added

# Database configuration
DATABASE = 'neural_log.db'

def get_db_connection():
    """Create a database connection"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize the database with required tables"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Create users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            email TEXT,
            is_admin INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create activities table with user_id
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS activities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            activity_name TEXT NOT NULL,
            description TEXT,
            duration INTEGER,
            progress_score INTEGER,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    # Create milestones table with user_id
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS milestones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            milestone_day INTEGER NOT NULL,
            insights TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    conn.commit()
    conn.close()

# Decorator for routes that require login
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorator for routes that require admin access
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        conn = get_db_connection()
        user = conn.execute('SELECT is_admin FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        conn.close()
        
        if not user or not user['is_admin']:
            return jsonify({'error': 'Admin access required'}), 403
        
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle user login"""
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['is_admin'] = bool(user['is_admin'])
            return jsonify({'success': True, 'username': user['username'], 'is_admin': bool(user['is_admin'])})
        else:
            return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Handle user registration"""
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        email = data.get('email', '')
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password required'}), 400
        
        conn = get_db_connection()
        
        # Check if username already exists
        existing_user = conn.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
        if existing_user:
            conn.close()
            return jsonify({'success': False, 'message': 'Username already exists'}), 400
        
        # Check if this is the first user (make them admin)
        user_count = conn.execute('SELECT COUNT(*) as count FROM users').fetchone()['count']
        is_admin = 1 if user_count == 0 else 0
        
        # Create new user
        password_hash = generate_password_hash(password)
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO users (username, password_hash, email, is_admin) VALUES (?, ?, ?, ?)',
            (username, password_hash, email, is_admin)
        )
        conn.commit()
        user_id = cursor.lastrowid
        conn.close()
        
        # Log in the new user
        session['user_id'] = user_id
        session['username'] = username
        session['is_admin'] = bool(is_admin)
        
        return jsonify({'success': True, 'username': username, 'is_admin': bool(is_admin)}), 201
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Handle user logout"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/current-user')
@login_required
def current_user():
    """Get current logged in user"""
    return jsonify({
        'user_id': session.get('user_id'),
        'username': session.get('username'),
        'is_admin': session.get('is_admin', False)
    })

@app.route('/')
@login_required
def index():
    """Render the main page"""
    return render_template('index.html')

@app.route('/api/activities', methods=['GET', 'POST'])
@login_required
def activities():
    """Handle GET and POST requests for activities"""
    user_id = session.get('user_id')
    conn = get_db_connection()
    
    if request.method == 'POST':
        data = request.json
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO activities (user_id, date, activity_name, description, duration, progress_score, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            user_id,
            data['date'],
            data['activity_name'],
            data.get('description', ''),
            data.get('duration', 0),
            data.get('progress_score', 0),
            data.get('notes', '')
        ))
        conn.commit()
        activity_id = cursor.lastrowid
        conn.close()
        return jsonify({'success': True, 'id': activity_id}), 201
    
    # GET request - only return activities for current user
    activities = conn.execute(
        'SELECT * FROM activities WHERE user_id = ? ORDER BY date DESC',
        (user_id,)
    ).fetchall()
    conn.close()
    
    return jsonify([dict(row) for row in activities])

@app.route('/api/activities/<int:activity_id>', methods=['DELETE'])
@login_required
def delete_activity(activity_id):
    """Delete a specific activity"""
    user_id = session.get('user_id')
    conn = get_db_connection()
    conn.execute('DELETE FROM activities WHERE id = ? AND user_id = ?', (activity_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/stats')
@login_required
def stats():
    """Get statistics about activities"""
    user_id = session.get('user_id')
    conn = get_db_connection()
    
    # Get total days logged
    total_days = conn.execute('''
        SELECT COUNT(DISTINCT date) as count FROM activities WHERE user_id = ?
    ''', (user_id,)).fetchone()['count']
    
    # Get total activities
    total_activities = conn.execute('''
        SELECT COUNT(*) as count FROM activities WHERE user_id = ?
    ''', (user_id,)).fetchone()['count']
    
    # Get average progress score
    avg_score = conn.execute('''
        SELECT AVG(progress_score) as avg FROM activities WHERE user_id = ? AND progress_score > 0
    ''', (user_id,)).fetchone()['avg'] or 0
    
    # Get activities by date for chart
    activities_by_date = conn.execute('''
        SELECT date, COUNT(*) as count, AVG(progress_score) as avg_score
        FROM activities
        WHERE user_id = ?
        GROUP BY date
        ORDER BY date
    ''', (user_id,)).fetchall()
    
    conn.close()
    
    return jsonify({
        'total_days': total_days,
        'total_activities': total_activities,
        'avg_score': round(avg_score, 2),
        'activities_by_date': [dict(row) for row in activities_by_date]
    })

@app.route('/api/milestones/<int:days>')
@login_required
def get_milestone(days):
    """Get milestone insights for specific day count"""
    if days not in [10, 25, 45, 70, 100]:
        return jsonify({'error': 'Invalid milestone day'}), 400
    
    user_id = session.get('user_id')
    conn = get_db_connection()
    
    # Get all activities up to this milestone for current user
    activities = conn.execute('''
        SELECT * FROM activities WHERE user_id = ? ORDER BY date
    ''', (user_id,)).fetchall()
    
    # Calculate insights
    total_activities = len(activities)
    avg_score = sum(a['progress_score'] for a in activities if a['progress_score']) / max(total_activities, 1)
    unique_activity_types = len(set(a['activity_name'] for a in activities))
    total_duration = sum(a['duration'] for a in activities if a['duration'])
    
    # Get activity distribution
    activity_counts = {}
    for activity in activities:
        name = activity['activity_name']
        activity_counts[name] = activity_counts.get(name, 0) + 1
    
    insights = {
        'milestone_day': days,
        'total_activities': total_activities,
        'avg_progress_score': round(avg_score, 2),
        'unique_activity_types': unique_activity_types,
        'total_duration_hours': round(total_duration / 60, 2),
        'activity_distribution': activity_counts,
        'activities': [dict(row) for row in activities]
    }
    
    # Save milestone
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO milestones (user_id, milestone_day, insights)
        VALUES (?, ?, ?)
    ''', (user_id, days, json.dumps(insights)))
    conn.commit()
    conn.close()
    
    return jsonify(insights)

@app.route('/api/export/excel')
@login_required
def export_excel():
    """Export all activities to Excel file"""
    user_id = session.get('user_id')
    username = session.get('username')
    conn = get_db_connection()
    activities = conn.execute(
        'SELECT * FROM activities WHERE user_id = ? ORDER BY date',
        (user_id,)
    ).fetchall()
    conn.close()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activities"
    
    # Headers
    headers = ['Date', 'Activity Name', 'Description', 'Duration (min)', 'Progress Score', 'Notes']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for activity in activities:
        ws.append([
            activity['date'],
            activity['activity_name'],
            activity['description'],
            activity['duration'],
            activity['progress_score'],
            activity['notes']
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    filename = f'neural_log_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join('exports', filename)
    
    # Create exports directory if it doesn't exist
    os.makedirs('exports', exist_ok=True)
    
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# Admin Routes
@app.route('/admin')
@admin_required
def admin_dashboard():
    """Admin dashboard page"""
    return render_template('admin.html')

@app.route('/api/admin/users')
@admin_required
def get_all_users():
    """Get all users (admin only)"""
    conn = get_db_connection()
    users = conn.execute('''
        SELECT id, username, email, is_admin, created_at,
               (SELECT COUNT(*) FROM activities WHERE user_id = users.id) as activity_count
        FROM users
        ORDER BY created_at DESC
    ''').fetchall()
    conn.close()
    
    return jsonify([dict(row) for row in users])

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    """Delete a user and all their data (admin only)"""
    # Prevent deleting yourself
    if user_id == session.get('user_id'):
        return jsonify({'success': False, 'message': 'Cannot delete your own account'}), 400
    
    conn = get_db_connection()
    
    # Delete user's activities and milestones
    conn.execute('DELETE FROM activities WHERE user_id = ?', (user_id,))
    conn.execute('DELETE FROM milestones WHERE user_id = ?', (user_id,))
    
    # Delete user
    conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/admin/users/<int:user_id>/toggle-admin', methods=['POST'])
@admin_required
def toggle_admin(user_id):
    """Toggle admin status for a user (admin only)"""
    # Prevent removing your own admin status
    if user_id == session.get('user_id'):
        return jsonify({'success': False, 'message': 'Cannot modify your own admin status'}), 400
    
    conn = get_db_connection()
    user = conn.execute('SELECT is_admin FROM users WHERE id = ?', (user_id,)).fetchone()
    
    if not user:
        conn.close()
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    new_admin_status = 0 if user['is_admin'] else 1
    conn.execute('UPDATE users SET is_admin = ? WHERE id = ?', (new_admin_status, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'is_admin': bool(new_admin_status)})

@app.route('/api/admin/users/<int:user_id>/activities')
@admin_required
def get_user_activities(user_id):
    """Get all activities for a specific user (admin only)"""
    conn = get_db_connection()
    activities = conn.execute('''
        SELECT * FROM activities WHERE user_id = ? ORDER BY date DESC
    ''', (user_id,)).fetchall()
    conn.close()
    
    return jsonify([dict(row) for row in activities])

@app.route('/api/admin/stats')
@admin_required
def admin_stats():
    """Get overall system statistics (admin only)"""
    conn = get_db_connection()
    
    total_users = conn.execute('SELECT COUNT(*) as count FROM users').fetchone()['count']
    total_activities = conn.execute('SELECT COUNT(*) as count FROM activities').fetchone()['count']
    total_admins = conn.execute('SELECT COUNT(*) as count FROM users WHERE is_admin = 1').fetchone()['count']
    
    # Most active users
    active_users = conn.execute('''
        SELECT users.username, COUNT(activities.id) as activity_count
        FROM users
        LEFT JOIN activities ON users.id = activities.user_id
        GROUP BY users.id
        ORDER BY activity_count DESC
        LIMIT 10
    ''').fetchall()
    
    conn.close()
    
    return jsonify({
        'total_users': total_users,
        'total_activities': total_activities,
        'total_admins': total_admins,
        'active_users': [dict(row) for row in active_users]
    })

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)

